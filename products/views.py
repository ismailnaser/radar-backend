import logging
import traceback

from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, status, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Product, ProductGalleryImage, SponsoredAd, Subscription, Favorite, StoreFavorite, SubscriptionRenewalRequest
from .serializers import (
    ProductSerializer,
    SponsoredAdSerializer,
    SubscriptionSerializer,
    FavoriteSerializer,
    StoreFavoriteSerializer,
    SubscriptionRenewalRequestSerializer,
    AdminAppPaymentSerializer,
)
from stores.models import CommunityServicePoint, StoreProfile
from stores.subscription_visibility import (
    ensure_subscription_for_store,
    queryset_public_stores_only,
    store_is_publicly_visible,
)
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.db.models import Q, Sum
from rest_framework.exceptions import PermissionDenied

from .ad_lifecycle import purge_expired_sponsored_ads
from common.image_webp import image_file_to_webp_content

logger = logging.getLogger(__name__)


class MerchantRequiredPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.user_type == 'merchant'


def user_is_app_admin(user):
    if not user or not user.is_authenticated:
        return False
    return bool(
        getattr(user, 'user_type', None) == 'admin'
        or user.is_superuser
        or user.is_staff
    )


class AdminRequiredPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return user_is_app_admin(request.user)

class MerchantProductListCreateView(generics.ListCreateAPIView):
    serializer_class = ProductSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_queryset(self):
        return Product.objects.filter(store__user=self.request.user).prefetch_related('gallery_images')

    def perform_create(self, serializer):
        store = StoreProfile.objects.get(user=self.request.user)
        serializer.save(store=store)


class MerchantProductExportExcelView(APIView):
    """تصدير جميع منتجات التاجر الحالي إلى ملف Excel."""
    permission_classes = [MerchantRequiredPermission]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        import json

        store = StoreProfile.objects.get(user=request.user)
        products = Product.objects.filter(store=store).order_by('id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'منتجاتي'
        ws.sheet_view.rightToLeft = True

        # Header styling
        header_font = Font(name='Calibri', bold=True, size=12, color='1A1D26')
        header_fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E8E6E0'),
            right=Side(style='thin', color='E8E6E0'),
            top=Side(style='thin', color='E8E6E0'),
            bottom=Side(style='thin', color='E8E6E0'),
        )

        headers = ['اسم المنتج', 'السعر', 'وصف المنتج', 'تفاصيل المنتج', 'الحالة']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12

        data_font = Font(name='Calibri', size=11)
        data_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        for row_num, product in enumerate(products, 2):
            features_str = ''
            if product.product_features:
                features_str = ' | '.join(product.product_features) if isinstance(product.product_features, list) else str(product.product_features)

            row_data = [
                product.name,
                float(product.price),
                product.description or '',
                features_str,
                'نشط' if not product.is_archived else 'مؤرشف',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        safe_name = (store.store_name or 'products').replace(' ', '_')[:30]
        response['Content-Disposition'] = f'attachment; filename="radar_products_{safe_name}.xlsx"'
        wb.save(response)
        return response


class MerchantProductImportExcelView(APIView):
    """استيراد منتجات من ملف Excel — يُنشئ منتجات جديدة من كل صف."""
    permission_classes = [MerchantRequiredPermission]

    def post(self, request):
        import openpyxl
        import os
        from io import BytesIO

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response(
                {'error': 'يرجى رفع ملف Excel (.xlsx)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'صيغة الملف غير مدعومة. الرجاء رفع ملف بصيغة .xlsx'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(BytesIO(excel_file.read()), read_only=True)
        except Exception:
            return Response(
                {'error': 'تعذر قراءة الملف. تأكد أنه ملف Excel صالح (.xlsx)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        store = StoreProfile.objects.get(user=request.user)
        uploaded_images = request.FILES.getlist('images')
        image_by_name = {}
        for f in uploaded_images:
            # Case-insensitive filename matching (including extension case, e.g. IMG1.JPG == img1.jpg)
            nm = os.path.basename(str(getattr(f, 'name', '') or '')).strip().casefold()
            if nm and nm not in image_by_name:
                image_by_name[nm] = f

        # Detect header row
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response(
                {'error': 'الملف فارغ أو يحتوي فقط على صف العناوين. أضف منتجات بدءاً من الصف الثاني.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        skipped_rows = []
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            # Skip completely empty rows
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Extract cell values (columns: name, price, description, features, images)
            name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            price_raw = row[1] if len(row) > 1 and row[1] is not None else ''
            description = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            features_raw = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            images_raw = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''

            if not name:
                skipped_rows.append(f'صف {row_idx}: اسم المنتج مطلوب')
                continue

            # Parse price
            try:
                price_val = float(str(price_raw).replace(',', '').replace('₪', '').strip())
                if price_val < 0:
                    raise ValueError()
            except (ValueError, TypeError):
                skipped_rows.append(f'صف {row_idx}: السعر غير صالح ({price_raw})')
                continue

            # Parse features
            features_list = []
            if features_raw:
                features_list = [f.strip() for f in features_raw.split('|') if f.strip()][:5]

            # Parse image names (optional): match against uploaded files by filename.
            matched_image_files = []
            missing_image_names = []
            if images_raw:
                seen = set()
                for raw_part in images_raw.split('|'):
                    part = raw_part.strip()
                    if not part:
                        continue
                    key = os.path.basename(part).strip().casefold()
                    if not key or key in seen:
                        continue
                    seen.add(key)
                    f = image_by_name.get(key)
                    if f is not None:
                        matched_image_files.append(f)
                    else:
                        missing_image_names.append(part)
                matched_image_files = matched_image_files[:5]
                if missing_image_names:
                    skipped_rows.append(
                        f"صف {row_idx}: بعض الصور غير موجودة ضمن الملفات المرفوعة ({', '.join(missing_image_names[:3])})"
                    )

            try:
                product = Product.objects.create(
                    store=store,
                    name=name,
                    price=price_val,
                    description=description,
                    product_features=features_list,
                    is_archived=False,
                )
                if matched_image_files:
                    for i, img_file in enumerate(matched_image_files):
                        # Convert to WebP before linking to the product gallery.
                        webp_file = image_file_to_webp_content(img_file)
                        final_file = webp_file or img_file
                        ProductGalleryImage.objects.create(
                            product=product,
                            image=final_file,
                            sort_order=i,
                        )
                created_count += 1
            except Exception as e:
                errors.append(f'صف {row_idx}: {str(e)[:60]}')

        result = {
            'message': f'تمت إضافة {created_count} منتج بنجاح.',
            'created_count': created_count,
        }
        if skipped_rows:
            result['skipped'] = skipped_rows
        if errors:
            result['errors'] = errors
        return Response(result, status=status.HTTP_201_CREATED if created_count > 0 else status.HTTP_200_OK)

class MerchantProductUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ProductSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_queryset(self):
        return Product.objects.filter(store__user=self.request.user).prefetch_related('gallery_images')

class AdRequestView(generics.ListCreateAPIView):
    serializer_class = SponsoredAdSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_queryset(self):
        return (
            SponsoredAd.objects.filter(store__user=self.request.user)
            .prefetch_related('gallery_images')
            .order_by('-created_at')
        )

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx

    def create(self, request, *args, **kwargs):
        """يلتقط أي استثناء أثناء إنشاء طلب الإعلان ويطبع التفاصيل كاملة في السجلات (stdout + logger)."""
        try:
            return super().create(request, *args, **kwargs)
        except Exception as exc:
            err_line = f"[POST merchant/ads] {type(exc).__name__}: {exc}"
            print(err_line, flush=True)
            print(traceback.format_exc(), flush=True)
            logger.exception("POST /api/products/merchant/ads/ failed: %s", err_line)
            raise

    def perform_create(self, serializer):
        store = StoreProfile.objects.get(user=self.request.user)
        ad = serializer.save(store=store)
        try:
            from users.models import AdminNotificationEvent

            AdminNotificationEvent.objects.create(
                event_type=AdminNotificationEvent.TYPE_AD_REQUEST,
                title="طلب إعلان ممول جديد",
                body=f"المتجر: {getattr(store, 'store_name', '—')}",
                created_by=self.request.user,
                related_app="products",
                related_id=getattr(ad, "id", None),
            )
        except Exception:
            pass


class MerchantAdUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SponsoredAdSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_queryset(self):
        return SponsoredAd.objects.filter(store__user=self.request.user)

    def update(self, request, *args, **kwargs):
        ad = self.get_object()
        if ad.status != 'pending':
            return Response({"error": "Cannot edit ad unless pending"}, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        ad = self.get_object()
        if ad.status != 'pending':
            return Response({"error": "Cannot delete ad unless pending"}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)


class AdminSponsoredAdListView(generics.ListAPIView):
    """كل الإعلانات الممولة (للمراجعة) — ?status=pending|active|rejected"""
    serializer_class = SponsoredAdSerializer
    permission_classes = [AdminRequiredPermission]

    def get_queryset(self):
        qs = SponsoredAd.objects.select_related("store").order_by("-created_at")
        st = self.request.query_params.get("status")
        if st in ("pending", "active", "rejected", "expired"):
            qs = qs.filter(status=st)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


class AdminSponsoredAdDetailView(generics.RetrieveAPIView):
    """تفاصيل إعلان واحد للمراجعة (إشعار الدفع داخل التطبيق)."""
    serializer_class = SponsoredAdSerializer
    permission_classes = [AdminRequiredPermission]
    queryset = SponsoredAd.objects.select_related("store")

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


class AdminPendingCountsView(APIView):
    """عدد الطلبات المعلّقة: إعلانات ممولة + تجديد اشتراك + نقاط خدمات مجتمعية — لشارات لوحة المدير."""

    permission_classes = [AdminRequiredPermission]

    def get(self, request):
        pending_ads = SponsoredAd.objects.filter(status="pending").count()
        pending_renewals = SubscriptionRenewalRequest.objects.filter(status="pending").count()
        pending_community_points = CommunityServicePoint.objects.filter(
            status=CommunityServicePoint.STATUS_PENDING
        ).count()
        pending_total = pending_ads + pending_renewals + pending_community_points
        return Response(
            {
                "pending_ads": pending_ads,
                "pending_renewals": pending_renewals,
                "pending_community_points": pending_community_points,
                "pending_total": pending_total,
            }
        )


class AdminSubscriptionRenewalListView(generics.ListAPIView):
    """طلبات تجديد الاشتراك — ?status=pending|approved|rejected"""
    serializer_class = SubscriptionRenewalRequestSerializer
    permission_classes = [AdminRequiredPermission]

    def get_queryset(self):
        qs = SubscriptionRenewalRequest.objects.select_related("store", "decided_by").order_by("-created_at")
        st = self.request.query_params.get("status")
        if st in ("pending", "approved", "rejected"):
            qs = qs.filter(status=st)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


class AdminAdSetStatusView(generics.GenericAPIView):
    permission_classes = [AdminRequiredPermission]

    def post(self, request, pk):
        ad = get_object_or_404(SponsoredAd, pk=pk)
        new_status = request.data.get("status")
        if new_status not in ("pending", "active", "rejected"):
            return Response({"error": "Invalid status"}, status=status.HTTP_400_BAD_REQUEST)
        ad.status = new_status
        if new_status == "active":
            ad.approved_at = timezone.now()
        else:
            ad.approved_at = None
        ad.save(update_fields=["status", "approved_at"])
        # سجل أرباح الإعلان عند تفعيله لأول مرة
        if new_status == "active":
            from .models import FinanceTransfer

            FinanceTransfer.objects.get_or_create(
                sponsored_ad=ad,
                defaults={
                    "kind": FinanceTransfer.KIND_AD,
                    "store": ad.store,
                    "payment_method": ad.payment_method,
                    "amount_ils": Decimal("5.00"),
                },
            )
        return Response({"message": "updated", "status": ad.status, "approved_at": ad.approved_at})

class SubscriptionStatusView(generics.RetrieveAPIView):
    serializer_class = SubscriptionSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_object(self):
        store = StoreProfile.objects.get(user=self.request.user)
        return ensure_subscription_for_store(store)


class MerchantSubscriptionRenewalRequestListCreateView(generics.ListCreateAPIView):
    serializer_class = SubscriptionRenewalRequestSerializer
    permission_classes = [MerchantRequiredPermission]

    def get_queryset(self):
        return SubscriptionRenewalRequest.objects.filter(store__user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        store = StoreProfile.objects.get(user=self.request.user)
        # رسوم تجديد الاشتراك ثابتة (10 شيكل) — تُستخدم في تقارير الأرباح.
        req = serializer.save(store=store, amount_ils=Decimal("10.00"))
        try:
            from users.models import AdminNotificationEvent

            AdminNotificationEvent.objects.create(
                event_type=AdminNotificationEvent.TYPE_SUBSCRIPTION_RENEWAL,
                title="طلب تجديد اشتراك جديد",
                body=f"المتجر: {getattr(store, 'store_name', '—')}",
                created_by=self.request.user,
                related_app="products",
                related_id=getattr(req, "id", None),
            )
        except Exception:
            pass


class AdminSubscriptionRenewalRequestApproveView(generics.GenericAPIView):
    permission_classes = [AdminRequiredPermission]

    def post(self, request, pk):
        req = get_object_or_404(SubscriptionRenewalRequest, pk=pk)
        if req.status != 'pending':
            return Response({"error": "Request already decided"}, status=status.HTTP_400_BAD_REQUEST)
        req.approve(decided_by=request.user, days=30)
        # سجل التحويل (إن لم يُسجّل)
        from .models import FinanceTransfer

        FinanceTransfer.objects.get_or_create(
            subscription_renewal=req,
            defaults={
                "kind": FinanceTransfer.KIND_SUBSCRIPTION_RENEWAL,
                "store": req.store,
                "payment_method": req.payment_method,
                "amount_ils": req.amount_ils or Decimal("10.00"),
            },
        )
        return Response({"message": "approved"})


class AdminSubscriptionRenewalRequestRejectView(generics.GenericAPIView):
    permission_classes = [AdminRequiredPermission]

    def post(self, request, pk):
        req = get_object_or_404(SubscriptionRenewalRequest, pk=pk)
        if req.status != 'pending':
            return Response({"error": "Request already decided"}, status=status.HTTP_400_BAD_REQUEST)
        req.reject(decided_by=request.user)
        return Response({"message": "rejected"})

class PublicProductListView(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        base = Product.objects.filter(is_archived=False).select_related('store')
        store_id = self.request.query_params.get('store_id')
        if store_id is not None and str(store_id).strip() != '':
            try:
                sid = int(store_id)
            except (TypeError, ValueError):
                return Product.objects.none()
            store = StoreProfile.objects.filter(pk=sid).first()
            if not store or not store_is_publicly_visible(store):
                return Product.objects.none()
            return base.filter(store_id=sid)
        visible_ids = queryset_public_stores_only(StoreProfile.objects.all()).values_list('pk', flat=True)
        return base.filter(store_id__in=visible_ids)

class PublicAdListView(generics.ListAPIView):
    serializer_class = SponsoredAdSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        purge_expired_sponsored_ads()
        cutoff = timezone.now() - timedelta(hours=24)
        visible_stores = queryset_public_stores_only(StoreProfile.objects.all())
        qs = (
            SponsoredAd.objects.filter(
                status='active',
                approved_at__isnull=False,
                approved_at__gte=cutoff,
                store__in=visible_stores,
            )
            .select_related('store', 'store__category', 'product')
            .prefetch_related('gallery_images')
            .order_by('-approved_at')
        )
        raw_cat = self.request.query_params.get('category')
        if raw_cat is not None and str(raw_cat).strip() != '':
            parts = [p.strip() for p in str(raw_cat).split(',') if p.strip()]
            ids = []
            for p in parts:
                try:
                    cid = int(p)
                    if cid > 0:
                        ids.append(cid)
                except (TypeError, ValueError):
                    continue
            if len(ids) == 1:
                qs = qs.filter(store__category_id=ids[0])
            elif len(ids) > 1:
                qs = qs.filter(store__category_id__in=ids)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

class FavoriteViewSet(viewsets.ModelViewSet):
    serializer_class = FavoriteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return (
            Favorite.objects.filter(user=self.request.user)
            .select_related('product', 'product__store', 'sponsored_ad', 'sponsored_ad__store')
            .prefetch_related('product__gallery_images', 'sponsored_ad__gallery_images')
        )

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        product = serializer.validated_data.get('product')
        sponsored_ad = serializer.validated_data.get('sponsored_ad')
        if product is None and sponsored_ad is not None and sponsored_ad.product_id is None:
            fav = Favorite.objects.filter(
                user=request.user,
                product__isnull=True,
                sponsored_ad=sponsored_ad,
            ).first()
            if fav:
                created = False
            else:
                fav = Favorite.objects.create(
                    user=request.user,
                    product=None,
                    sponsored_ad=sponsored_ad,
                )
                created = True
        else:
            fav, created = Favorite.objects.get_or_create(
                user=request.user,
                product=product,
                defaults={'sponsored_ad': sponsored_ad},
            )
            if not created and sponsored_ad is not None:
                fav.sponsored_ad = sponsored_ad
                fav.save(update_fields=['sponsored_ad'])
        out = FavoriteSerializer(fav, context=self.get_serializer_context())
        return Response(out.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class StoreFavoriteViewSet(viewsets.ModelViewSet):
    serializer_class = StoreFavoriteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return StoreFavorite.objects.filter(user=self.request.user).select_related('store', 'store__category')

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


def user_is_primary_admin(user):
    if not user or not user.is_authenticated:
        return False
    return bool(getattr(user, "is_primary_admin", False) or user.is_superuser)


class PrimaryAdminRequiredPermission(permissions.BasePermission):
    message = "هذا الإجراء متاح للمدير الأساسي فقط."

    def has_permission(self, request, view):
        return user_is_primary_admin(request.user)


class AdminFinanceTransfersView(APIView):
    """تقارير الأرباح/التحويلات — للمدير الأساسي فقط.

    Query params:
      - q: بحث باسم المتجر أو اسم المستخدم أو الجوال
      - from: YYYY-MM-DD
      - to: YYYY-MM-DD
      - method: balipay_wallet|bank_palestine|other
      - kind: sponsored_ad|subscription_renewal
    """

    permission_classes = [PrimaryAdminRequiredPermission]

    def get(self, request):
        from .models import FinanceTransfer

        def abs_media(u):
            if not u:
                return None
            try:
                url = request.build_absolute_uri(u)
                xf_proto = (request.META.get("HTTP_X_FORWARDED_PROTO") or "").split(",")[0].strip().lower()
                if xf_proto == "https" and url.startswith("http://"):
                    url = "https://" + url[len("http://") :]
                return url
            except Exception:
                return u

        qs = (
            FinanceTransfer.objects.select_related("store", "store__user")
            .order_by("-created_at")
        )

        q = (request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(store__store_name__icontains=q)
                | Q(store__user__username__icontains=q)
                | Q(store__user__phone_number__icontains=q)
            )

        method = (request.query_params.get("method") or "").strip()
        if method in ("balipay_wallet", "bank_palestine", "other"):
            qs = qs.filter(payment_method=method)

        kind = (request.query_params.get("kind") or "").strip()
        if kind in ("sponsored_ad", "subscription_renewal"):
            qs = qs.filter(kind=kind)

        from_raw = (request.query_params.get("from") or "").strip()
        to_raw = (request.query_params.get("to") or "").strip()
        # التاريخ باليوم (حسب created_at)
        if from_raw:
            try:
                qs = qs.filter(created_at__date__gte=from_raw)
            except Exception:
                pass
        if to_raw:
            try:
                qs = qs.filter(created_at__date__lte=to_raw)
            except Exception:
                pass

        total_amount = qs.aggregate(total=Sum("amount_ils")).get("total") or Decimal("0.00")
        total_count = qs.count()

        results = []
        for t in qs[:500]:
            u = getattr(t.store, "user", None)
            receipt_url = None
            try:
                if getattr(t, "sponsored_ad_id", None) and getattr(t, "sponsored_ad", None):
                    img = getattr(t.sponsored_ad, "payment_receipt_image", None)
                    if img:
                        receipt_url = abs_media(img.url)
                elif getattr(t, "subscription_renewal_id", None) and getattr(t, "subscription_renewal", None):
                    img = getattr(t.subscription_renewal, "receipt_image", None)
                    if img:
                        receipt_url = abs_media(img.url)
            except Exception:
                receipt_url = None
            results.append(
                {
                    "id": t.id,
                    "kind": t.kind,
                    "kind_label": t.get_kind_display(),
                    "store_id": t.store_id,
                    "store_name": t.store.store_name,
                    "merchant_username": getattr(u, "username", None),
                    "merchant_phone": getattr(u, "phone_number", None),
                    "payment_method": t.payment_method,
                    "payment_method_label": t.get_payment_method_display(),
                    "amount_ils": str(t.amount_ils),
                    "created_at": t.created_at,
                    "receipt_image": receipt_url,
                }
            )

        return Response(
            {
                "meta": {
                    "total_count": total_count,
                    "total_amount_ils": str(total_amount),
                },
                "results": results,
            }
        )


class AdminFinanceTransferDetailView(APIView):
    """حذف تحويلة أرباح/تحويلات محددة — للمدير الأساسي فقط."""

    permission_classes = [PrimaryAdminRequiredPermission]

    def delete(self, request, transfer_id: int):
        from .models import FinanceTransfer

        t = FinanceTransfer.objects.filter(pk=transfer_id).first()
        if not t:
            return Response({"detail": "غير موجود."}, status=status.HTTP_404_NOT_FOUND)
        t.delete()
        return Response({"ok": True})


class AdminAppPaymentsView(APIView):
    """مدفوعات الإدارة للتطبيق (مدفوع/قيد الدفع) — للمدير الأساسي فقط."""

    permission_classes = [PrimaryAdminRequiredPermission]

    def get(self, request):
        from .models import AdminAppPayment
        qs = AdminAppPayment.objects.select_related("created_by").all()
        status_q = (request.query_params.get("status") or "").strip()
        if status_q in ("paid", "planned"):
            qs = qs.filter(status=status_q)
        total_paid = qs.filter(status=AdminAppPayment.STATUS_PAID).aggregate(total=Sum("amount_ils")).get("total") or Decimal("0.00")
        total_all = qs.aggregate(total=Sum("amount_ils")).get("total") or Decimal("0.00")
        ser = AdminAppPaymentSerializer(qs[:500], many=True, context={"request": request})
        return Response(
            {
                "meta": {
                    "total_count": qs.count(),
                    "total_paid_ils": str(total_paid),
                    "total_all_ils": str(total_all),
                },
                "results": ser.data,
            }
        )

    def post(self, request):
        ser = AdminAppPaymentSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save(created_by=request.user)
        return Response(AdminAppPaymentSerializer(obj, context={"request": request}).data, status=status.HTTP_201_CREATED)


class AdminAppPaymentDetailView(APIView):
    permission_classes = [PrimaryAdminRequiredPermission]

    def patch(self, request, payment_id: int):
        from .models import AdminAppPayment
        obj = AdminAppPayment.objects.filter(pk=payment_id).first()
        if not obj:
            return Response({"detail": "غير موجود."}, status=status.HTTP_404_NOT_FOUND)
        ser = AdminAppPaymentSerializer(obj, data=request.data, partial=True, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    def delete(self, request, payment_id: int):
        from .models import AdminAppPayment
        obj = AdminAppPayment.objects.filter(pk=payment_id).first()
        if not obj:
            return Response({"detail": "غير موجود."}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response({"ok": True})
